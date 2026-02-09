# SPDX-FileCopyrightText: 2026 TOP Team Combat Control
#
# SPDX-License-Identifier: GPL-3.0-or-later

"""Pytest smoke tests for all dependencies."""
from main import test_all_imports


def test_all_libraries_import():
    results = test_all_imports()
    failed = [(name, info) for name, ok, info in results if not ok]
    assert not failed, f"Failed imports: {failed}"


def test_fastapi_app():
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    app = FastAPI()

    @app.get("/health")
    def health():
        return {"status": "ok"}

    client = TestClient(app)
    resp = client.get("/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok"}


def test_openpyxl_create():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Score"])
    ws.append(["Test", 100])
    assert ws.cell(row=2, column=1).value == "Test"


def test_fpdf2_create():
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=16)
    pdf.cell(text="JudgeFrontend PDF Test")
    output = pdf.output()
    assert len(output) > 0


def test_opencv_basic():
    import cv2
    import numpy as np
    img = np.zeros((100, 100, 3), dtype=np.uint8)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    assert gray.shape == (100, 100)
